import os
from openpyxl import load_workbook, Workbook

# Function to copy sheet to new workbook
def copy_sheet(source_sheet, destination_workbook, new_sheet_name):
    new_sheet = destination_workbook.create_sheet(title=new_sheet_name)
    for row in source_sheet.iter_rows():
        for cell in row:
            new_sheet[cell.coordinate].value = cell.value

# Prompt to select sheet
def select_sheet(sheet_names):
    print("Available sheets:")
    for i, sheet_name in enumerate(sheet_names):
        print(f"{i + 1}. {sheet_name}")
    selection = input("Enter the number of the sheet you want to select: ")
    try:
        selection = int(selection)
        if 1 <= selection <= len(sheet_names):
            return sheet_names[selection - 1]
        else:
            print("Invalid selection. Please enter a valid sheet number.")
            return select_sheet(sheet_names)
    except ValueError:
        print("Invalid input. Please enter a number.")
        return select_sheet(sheet_names)

# Main function
def main():
    merged_workbook = Workbook()
    directories = ['directory1', 'directory2', 'directory3', 'directory4']
    for directory in directories:
        for file_name in os.listdir(directory):
            if file_name.endswith('.xlsx') and ("cam" in file_name.lower() or "san diego" in file_name.lower() or "uk" in file_name.lower()):
                file_path = os.path.join(directory, file_name)
                source_workbook = load_workbook(file_path)
                for sheet_name in source_workbook.sheetnames:
                    selected_sheet = select_sheet(source_workbook.sheetnames)
                    copy_sheet(source_workbook[sheet_name], merged_workbook, file_name[:20])
                    break  # Exit loop after processing one sheet
    merged_workbook.save("merged_data.xlsx")

if __name__ == "__main__":
    main()
